#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Conversor da base de cargos (planilha Excel) para o portal Wessel Carreiras.

Le a planilha em dados/Base_de_Cargos_Wessel.xlsx (aba "Base de Cargos"),
normaliza os campos e gera site/data.js com um objeto global window.WESSEL_DATA.

Campos vazios NAO sao inventados: ficam como string vazia e o portal os
exibe como "a definir". A ordem das colunas na planilha nao importa; o
mapeamento e feito por nome de coluna (tolerante a acentos e maiusculas).
"""

import json
import glob
import sys
import unicodedata
from datetime import datetime, date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl nao instalado. Rode: pip install openpyxl")

RAIZ = Path(__file__).resolve().parent.parent
PASTA_DADOS = RAIZ / "dados"
SAIDA = RAIZ / "site" / "data.js"
ABA = "Base de Cargos"


def sem_acento(texto):
    t = unicodedata.normalize("NFKD", str(texto)).encode("ascii", "ignore").decode()
    return " ".join(t.lower().split())


# chave interna -> lista de nomes de coluna aceitos (normalizados sem acento)
MAPA_COLUNAS = {
    "codigo":            ["codigo do cargo", "codigo"],
    "titulo":            ["titulo do cargo", "titulo"],
    "nivel":             ["nivel/grau", "nivel grau", "nivel"],
    "area":              ["area"],
    "setores":           ["setor(es) aplicavel(is)", "setores aplicaveis", "setor(es)", "setores"],
    "superior":          ["cargo superior imediato", "cargo superior"],
    "subordinados_diretos":   ["subordinados diretos"],
    "subordinados_indiretos": ["subordinados indiretos"],
    "subordinados":      ["subordinados diretos/indiretos", "subordinados"],
    "missao":            ["missao do cargo", "missao"],
    "responsabilidades": ["responsabilidades e atividades", "responsabilidades"],
    "escolaridade":      ["escolaridade"],
    "experiencia":       ["experiencia minima", "experiencia"],
    "comp_comportamentais": ["competencias comportamentais"],
    "comp_tecnicas":     ["competencias tecnicas"],
    "treinamentos":      ["treinamentos/nr obrigatorios", "treinamentos/nr", "treinamentos"],
    "requisitos_fisicos": ["requisitos fisicos/ambiente", "requisitos fisicos"],
    "viagens":           ["viagens/cnh/disponibilidade", "viagens/cnh", "viagens"],
    "outras":            ["outras observacoes", "observacoes"],
    "abas_origem":       ["abas de origem (arquivo antigo)", "abas de origem"],
    "elaborado_por":     ["elaborado por"],
    "data_criacao":      ["data de criacao"],
    "versao":            ["versao"],
    "validado_por":      ["validado por"],
    "status":            ["status"],
    "atualizacao":       ["ultima atualizacao"],
}

# campos que podem conter varios itens separados por ; ou quebra de linha
CAMPOS_LISTA = {
    "setores", "subordinados_diretos", "subordinados_indiretos", "subordinados",
    "comp_comportamentais", "comp_tecnicas", "treinamentos", "abas_origem",
}


def limpa(valor):
    """Normaliza um valor de celula para string limpa (ou '' se vazio)."""
    if valor is None:
        return ""
    if isinstance(valor, (datetime, date)):
        return valor.strftime("%d/%m/%Y")
    if isinstance(valor, float) and valor.is_integer():
        valor = int(valor)
    texto = str(valor).replace("\r\n", "\n").strip()
    marcadores_vazio = {"", "-", "n/a", "na", "a definir", "a definir."}
    if sem_acento(texto) in marcadores_vazio:
        return ""
    return texto


def como_lista(texto):
    """Quebra um campo multi-item em lista, aceitando ;, bullets e novas linhas."""
    if not texto:
        return []
    bruto = texto.replace("•", "\n").replace(";", "\n")
    itens = []
    for parte in bruto.split("\n"):
        p = parte.strip(" \t-–—")
        if p:
            itens.append(p)
    return itens


def acha_planilha():
    arquivos = sorted(glob.glob(str(PASTA_DADOS / "*.xlsx")))
    arquivos = [a for a in arquivos if not Path(a).name.startswith("~$")]
    if not arquivos:
        sys.exit(f"Nenhuma planilha .xlsx encontrada em {PASTA_DADOS}")
    if len(arquivos) > 1:
        print(f"Aviso: varias planilhas encontradas, usando a primeira: {arquivos[0]}")
    return arquivos[0]


def main():
    caminho = acha_planilha()
    print(f"Lendo: {caminho}")
    wb = openpyxl.load_workbook(caminho, data_only=True)
    if ABA not in wb.sheetnames:
        sys.exit(f'Aba "{ABA}" nao encontrada. Abas: {wb.sheetnames}')
    ws = wb[ABA]

    cabecalho = [limpa(c.value) for c in ws[1]]
    indice = {}
    for chave, nomes in MAPA_COLUNAS.items():
        for i, nome in enumerate(cabecalho):
            if sem_acento(nome) in nomes:
                indice[chave] = i
                break

    obrigatorias = ["codigo", "titulo", "area", "setores"]
    faltando = [c for c in obrigatorias if c not in indice]
    if faltando:
        sys.exit(f"Colunas obrigatorias ausentes na planilha: {faltando}")

    cargos = []
    for linha in ws.iter_rows(min_row=2, values_only=True):
        if not any(v not in (None, "") for v in linha):
            continue
        reg = {}
        for chave, i in indice.items():
            valor = limpa(linha[i]) if i < len(linha) else ""
            reg[chave] = como_lista(valor) if chave in CAMPOS_LISTA else valor
        # unifica subordinados diretos/indiretos numa lista, se vierem separados
        if "subordinados_diretos" in reg or "subordinados_indiretos" in reg:
            reg["subordinados"] = reg.get("subordinados_diretos", []) + reg.get("subordinados_indiretos", [])
        if not reg.get("codigo") and not reg.get("titulo"):
            continue
        cargos.append(reg)

    cargos.sort(key=lambda r: (r.get("area", ""), r.get("titulo", "")))

    # indice de areas -> setores -> cargos (para navegacao)
    areas = {}
    for reg in cargos:
        for area in [a.strip() for a in reg.get("area", "").split(";") if a.strip()] or ["Sem area"]:
            areas.setdefault(area, set())
            for s in reg.get("setores", []):
                areas[area].add(s)
    areas_ordenadas = {a: sorted(areas[a]) for a in sorted(areas)}

    dados = {
        "portal": "Wessel Carreiras",
        "geradoEm": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "totalCargos": len(cargos),
        "areas": areas_ordenadas,
        "cargos": cargos,
    }

    SAIDA.parent.mkdir(parents=True, exist_ok=True)
    corpo = json.dumps(dados, ensure_ascii=False, separators=(",", ":"))
    SAIDA.write_text(
        "/* Gerado automaticamente por scripts/build.py. Nao editar a mao. */\n"
        "window.WESSEL_DATA = " + corpo + ";\n",
        encoding="utf-8",
    )
    print(f"OK: {len(cargos)} cargos, {len(areas_ordenadas)} areas -> {SAIDA}")


if __name__ == "__main__":
    main()
